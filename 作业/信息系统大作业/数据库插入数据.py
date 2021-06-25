import psycopg2
import openpyxl

conn = psycopg2.connect(host='localhost', 
	database="postgres", user="postgres", password="pass")

cur = conn.cursor()
cur.execute('''
DROP TABLE IF EXISTS course;
CREATE TABLE IF NOT EXISTS course  (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --课程号
    name     TEXT,        --课程名称
    teacher  TEXT,
    data     TEXT,        --课程时间
    PRIMARY KEY(sn));
''')

cur.execute('''
DROP TABLE IF EXISTS student;
CREATE TABLE IF NOT EXISTS student (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --学号
    name     TEXT,        --姓名
    clss    TEXT,        --班级
    PRIMARY KEY(sn));
''')


cur.execute('''
DROP TABLE IF EXISTS course_grade;
CREATE TABLE IF NOT EXISTS course_grade (
    stu_sn VARCHAR(20),     -- 学生序号
    cou_sn VARCHAR(20),     -- 课程序号
    grade  NUMERIC(5,2), -- 最终成绩
    PRIMARY KEY(stu_sn, cou_sn));
''')


wb=openpyxl.load_workbook("信息系统数据.xlsx")
ws=wb.active
colC=ws["C"]    #导入数据库列单元
colD=ws["D"]
colF=ws["F"]
colG=ws["G"]
colH=ws["H"]
colI=ws["I"]
co1J=ws["J"]
colL=ws["L"]
colM=ws["M"]
colN=ws["N"]

for i in range(1,21):
    sn = i
    no= '%s' % colG[i].value
    name = '%s' % colI[i].value
    teacher = '%s' % colH[i].value
    data = '%s' % co1J[i].value
    cur.execute('''
		INSERT INTO course(sn, no,teacher,name,data) VALUES (%(sn)s, %(no)s,%(teacher)s,%(name)s,%(data)s) 
	''', {'sn':sn, 'no':no,"teacher":teacher,'name':name,"data":data} )

for i in range(1,41):
    sn = i
    no=1910610000+i
    name = '%s' % colC[i].value
    clss = '%s' % colD[i].value
    cur.execute('''
		INSERT INTO student(sn, no,name,clss) VALUES (%(no)s, %(sn)s,%(name)s,%(clss)s) 
	''', {'sn':sn, 'no':no,'name':name,"clss":clss} )


for i in range(1,len(colN)):
    stu_sn = '%s' % colL[i].value
    cou_sn= '%s' % colM[i].value
    grade='%s' % colN[i].value
    cur.execute('''
		INSERT INTO course_grade(stu_sn,cou_sn,grade) VALUES (%(stu_sn)s,%(cou_sn)s,%(grade)s) 
	''', {'stu_sn':stu_sn, 'cou_sn':cou_sn,'grade':grade} )

conn.commit()